import xml.etree.ElementTree as ET
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import logging
from bs4 import BeautifulSoup
import re
import os
import openpyxl


# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ProPublicaScraper:
    """Scraper for ProPublica nonprofit search pages"""
    
    def __init__(self):
        self.base_url = "https://projects.propublica.org"
    
    def get_organization_links(self, main_url):
        """
        Extract XML download links and NTEE category from organization's main page
        Returns: tuple (ntee_category, list_of_xml_urls)
        """
        try:
            # Fetch the main page
            response = requests.get(main_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract NTEE category
            ntee_elem = soup.find('p', class_='ntee-category')
            if ntee_elem:
                ntee_category = ntee_elem.text.split(':', 1)[1].strip().split('/')[0].strip()
            else:
                ntee_category = "Unknown"
            
            # Find all XML download links
            xml_links = []
            for link in soup.find_all('a', class_='btn', target='_blank'):
                if 'XML' in link.text:
                    object_id = link['href'].split('object_id=')[1]
                    full_url = f"{self.base_url}/nonprofits/download-xml?object_id={object_id}"
                    xml_links.append(full_url)
            
            # Sort links by object_id (which contains year) and take most recent 5
            xml_links.sort(reverse=True)
            xml_links = xml_links[:5]
            
            return ntee_category, xml_links
            
        except requests.RequestException as e:
            logger.error(f"Error fetching organization page: {str(e)}")
            raise


class NonprofitParser:
    """Parser for nonprofit financial data from ProPublica URLs"""
    
    def __init__(self):
        # Initialize scraper and remove test_urls since we'll use dynamic scraping
        self.scraper = ProPublicaScraper()
        self.ns = {'irs': 'http://www.irs.gov/efile'}
        self.leadership_titles = [
            'PRESIDENT', 'CEO', 'CHIEF EXECUTIVE OFFICER',
            'CFO', 'CHIEF FINANCIAL OFFICER',
            'COO', 'CHIEF OPERATING OFFICER',
            'CHANCELLOR', 'DEAN',
            'TREASURER'
        ]

    def fetch_content(self, url):
        """Fetch content from URL with error handling"""
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response.content
        except requests.RequestException as e:
            logger.error(f"Failed to fetch from URL {url}: {str(e)}")
            raise

    def detect_format(self, content):
        """
        Detect whether content is XML or TXT format
        Returns tuple of (format_type, parsed_content)
        """
        try:
            # First try parsing as XML
            tree = ET.parse(BytesIO(content))
            logger.info("Successfully parsed as XML")
            return 'xml', tree
        except ET.ParseError:
            # If XML parsing fails, try TXT format
            try:
                text_content = content.decode('utf-8', errors='ignore')
                # Check for common TXT format markers
                if any(marker in text_content.upper() for marker in 
                      ['RETURN HEADER', 'FORM 990', 'EIN:']):
                    logger.info("Successfully parsed as TXT")
                    return 'txt', text_content
                else:
                    raise ValueError("Content doesn't match expected formats")
            except Exception as e:
                logger.error(f"Format detection failed: {str(e)}")
                raise

    def get_tax_year(self, content, format_type):
        """Extract tax year from content and adjust it to reflect reporting year"""
        try:
            if format_type == 'xml':
                root = content.getroot()
                # Try multiple possible locations for tax year
                tax_period = root.find('.//irs:TaxPeriodEndDt', self.ns)
                if tax_period is not None and tax_period.text:
                    # Subtract 1 from the tax year to get the reporting year
                    return str(int(datetime.strptime(tax_period.text, '%Y-%m-%d').year) - 1)
                
                tax_year = root.find('.//irs:TaxYr', self.ns)
                if tax_year is not None and tax_year.text:
                    # Subtract 1 from the tax year to get the reporting year
                    return str(int(tax_year.text) - 1)
            else:
                # Search for year in TXT content
                text_lines = content.split('\n')
                for line in text_lines:
                    if 'Tax Period Begin' in line:
                        # Extract first 4-digit number found and subtract 1
                        for word in line.split():
                            if word.isdigit() and len(word) == 4:
                                return str(int(word) - 1)
            
            logger.warning("Could not find tax year")
            return "Unknown"
        except Exception as e:
            logger.error(f"Error extracting tax year: {str(e)}")
            return "Unknown"

    def get_organization_name(self, content, format_type):
        """Extract and normalize organization name from content"""
        try:
            raw_name = None
            if format_type == 'xml':
                root = content.getroot()
                # Try multiple possible locations for organization name
                for path in [
                    './/irs:BusinessName/irs:BusinessNameLine1Txt',
                    './/irs:ReturnHeader/irs:Filer/irs:BusinessName/irs:BusinessNameLine1Txt'
                ]:
                    name = root.find(path, self.ns)
                    if name is not None and name.text:
                        raw_name = name.text
                        break
            else:
                # Search for organization name in TXT content
                text_lines = content.split('\n')
                for line in text_lines:
                    if 'Name of Organization:' in line or 'NAME OF ORGANIZATION:' in line:
                        raw_name = line.split(':', 1)[1].strip()
                        break
            
            if not raw_name:
                logger.warning("Could not find organization name")
                return "Unknown Organization"
                
            # Normalize the organization name to ensure consistency
            # 1. Convert to uppercase for case-insensitive comparison
            # 2. Remove common corporate suffixes
            # 3. Remove extra whitespace
            normalized_name = raw_name.upper()
            common_suffixes = [' INC', ' LLC', ' FOUNDATION', ' CORP', ' CORPORATION', ' LTD', ' INCORPORATED']
            for suffix in common_suffixes:
                if normalized_name.endswith(suffix):
                    normalized_name = normalized_name[:-len(suffix)]
            
            # Remove extra spaces and trim
            normalized_name = ' '.join(normalized_name.split())
            
            # Return the original name but use normalized version for matching
            return raw_name
            
        except Exception as e:
            logger.error(f"Error extracting organization name: {str(e)}")
            return "Unknown Organization"
    
    def process_url(self, url):
        """Process a single URL and return basic information"""
        try:
            content = self.fetch_content(url)
            format_type, parsed_content = self.detect_format(content)
            
            # Initialize financial data extractor
            extractor = FinancialDataExtractor()
            
            result = {
                'url': url,
                'format': format_type,
                'tax_year': self.get_tax_year(parsed_content, format_type),
                'organization_name': self.get_organization_name(parsed_content, format_type),
                'parsed_content': parsed_content,
            }
            
            # Extract financial metrics
            result['financial_metrics'] = extractor.extract_financial_metrics(
                parsed_content,
                format_type
            )
            
            # Extract endowment data
            result['endowment_data'] = extractor.extract_endowment_data(
                parsed_content,
                format_type
            )
            
            # Extract executive compensation
            result['executive_compensation'] = extractor.extract_executive_compensation(
                parsed_content,
                format_type
            )
        
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing URL {url}: {str(e)}")
            raise

class FinancialDataExtractor:
    """Extracts financial data from parsed nonprofit documents"""
    
    def __init__(self):
            self.ns = {'irs': 'http://www.irs.gov/efile'}
            # Add leadership titles here
            self.leadership_titles = [
                'PRESIDENT', 'CEO', 'CHIEF EXECUTIVE',
                'CFO', 'CHIEF FINANCIAL',
                'COO', 'CHIEF OPERATING',
                'CHANCELLOR', 'DEAN','TREASURER'
                # TODO: Add more leadership titles as needed
            ]
        
    def extract_financial_metrics(self, content, format_type):
        """Extract basic financial metrics"""
        try:
            if format_type == 'xml':
                return self._extract_financial_metrics_xml(content)
            else:
                return self._extract_financial_metrics_txt(content)
        except Exception as e:
            logger.error(f"Error extracting financial metrics: {str(e)}")
            return {}

    def _extract_financial_metrics_xml(self, tree):
        root = tree.getroot()
        metrics = {}

        # Handle TotalFunctionalExpensesGrp
        total_expenses = root.find('.//irs:TotalFunctionalExpensesGrp', self.ns)
        if total_expenses is not None:
            mgmt_total = total_expenses.find('.//irs:ManagementAndGeneralAmt', self.ns)
            fundraising_total = total_expenses.find('.//irs:FundraisingAmt', self.ns)
            if mgmt_total is not None:
                metrics['ManagementAndGeneralAmt'] = mgmt_total.text
            if fundraising_total is not None:
                metrics['CYTotalFundraisingExpenseAmt'] = fundraising_total.text

        # Handle group elements
        group_elements = {
            'InformationTechnologyGrp': 'InformationTechnologyGrp',
            'OccupancyGrp': 'OccupancyGrp',
            'TravelGrp': 'TravelGrp',
            'FeesForServicesAccountingGrp': 'FeesForServicesAccountingGrp'
        }

        for field, xml_tag in group_elements.items():
            group = root.find(f'.//irs:{xml_tag}', self.ns)
            if group is not None:
                total = group.find('.//irs:TotalAmt', self.ns)
                if total is not None:
                    metrics[field] = total.text
        
        donor_restriction_paths = {
            'WithoutDonorRestrictions': [
                ('.//irs:NoDonorRestrictionNetAssetsGrp/irs:EOYAmt', self.ns),
                ('.//irs:UnrestrictedNetAssetsGrp/irs:EOYAmt', self.ns)
            ],
            'WithDonorRestrictions': [
                ('.//irs:DonorRestrictionNetAssetsGrp/irs:EOYAmt', self.ns),
                ('.//irs:PermanentlyRstrNetAssetsGrp/irs:EOYAmt', self.ns)
            ]
        }
        
        for metric, paths in donor_restriction_paths.items():
            for path, ns in paths:
                value = root.find(path, ns)
                if value is not None and value.text:
                    metrics[metric] = value.text
                    break
            if metric not in metrics:
                metrics[metric] = 'Not found'
        
    
    
        balance_sheet_groups = {
            'CashNonInterestBearing': 'CashNonInterestBearingGrp',
            'AccountsReceivable': 'AccountsReceivableGrp',
            'AccountsPayable': 'AccountsPayableAccrExpnssGrp'
        }

        # Process each balance sheet group
        for field, group_name in balance_sheet_groups.items():
            group = root.find(f'.//irs:{group_name}', self.ns)
            if group is not None:
                eoy_amt = group.find('.//irs:EOYAmt', self.ns)
                if eoy_amt is not None:
                    metrics[f'{field}EOY'] = eoy_amt.text



        # Basic financial elements to extract
        financial_elements = {
            'revenue': [
                'CYTotalRevenueAmt',
                'CYContributionsGrantsAmt',
                'CYProgramServiceRevenueAmt',
                'InvestmentIncomeAmt',
                'CYOtherRevenueAmt',
                'CYInvestmentIncomeAmt',
                'CYRevenuesLessExpensesAmt'
            ],
            'expenses': [
                'CYTotalExpensesAmt',
                'CYGrantsAndSimilarPaidAmt',
                'CYSalariesCompEmpBnftPaidAmt',
                'TotalProgramServiceExpensesAmt',
                'FundraisingAmt',
                'CYOtherExpensesAmt',
                'OtherEmployeeBenefitsGrp/TotalAmt'
            ],
            'assets': [
                'TotalAssetsEOYAmt',
                'TotalLiabilitiesEOYAmt',
                'NetAssetsOrFundBalancesEOYAmt'
            ],

            'balance_sheet': [
                'CashNonInterestBearingGrp/EOYAmt',
                'AccountsReceivableGrp/EOYAmt',
                'AccountsPayableAccrExpnssGrp/EOYAmt'
            ],
            'other': [
                'TotalEmployeeCnt',
                'TotalVolunteersCnt'               
            ]
        }
        
        # Process regular financial elements
        for category, elements in financial_elements.items():
            for element in elements:
                paths = [
                    f'.//irs:{element}',
                    f'.//irs:IRS990/{element}',
                    f'.//irs:Form990PartIX/{element}'
                ]
                
                value = None
                for path in paths:
                    value = root.find(path, self.ns)
                    if value is not None:
                        break
                
                metrics[element] = value.text if value is not None else 'Not found'
                
        return metrics

    def _extract_financial_metrics_txt(self, content):
        """Extract basic financial metrics from TXT format"""
        metrics = {}
        lines = content.split('\n')
        
        # Add more comprehensive patterns for financial metrics
        field_patterns = {
            'CYTotalRevenueAmt': ['TOTAL REVENUE', 'REVENUE TOTAL'],
            'CYContributionsGrantsAmt': ['CONTRIBUTIONS AND GRANTS', 'GIFTS GRANTS', 'CONTRIBUTIONS GIFTS GRANTS', 'TOTAL CONTRIBUTIONS'],
            'CYProgramServiceRevenueAmt': ['PROGRAM SERVICE REVENUE', 'SERVICE REVENUE', 'PROGRAM REVENUE'],
            'CYInvestmentIncomeAmt': ['INVESTMENT INCOME', 'INVESTMENT EARNINGS', 'DIVIDENDS INTEREST'],
            'CYOtherRevenueAmt': ['OTHER REVENUE'],
            'CYTotalExpensesAmt': ['TOTAL EXPENSES', 'EXPENSES TOTAL'],
            'CYGrantsAndSimilarPaidAmt': ['GRANTS PAID', 'GRANTS AND SIMILAR AMOUNTS PAID'],
            'CYSalariesCompEmpBnftPaidAmt': ['SALARIES OTHER COMPENSATION', 'SALARIES AND WAGES', 'OFFICER COMPENSATION'],
            'TotalProgramServiceExpensesAmt': ['PROGRAM SERVICE EXPENSES', 'TOTAL PROGRAM SERVICE'],
            'FeesForServicesAccountingGrp': ['ACCOUNTING, ACCOUNTING FEE'],
            'ManagementAndGeneralAmt': ['MANAGEMENT AND GENERAL', 'MANAGEMENT EXPENSES'],
            'CYTotalFundraisingExpenseAmt': ['FUNDRAISING EXPENSES', 'FUNDRAISING COSTS', 'TOTAL FUNDRAISING'],
            'CYRevenuesLessExpensesAmt': ['REVENUE LESS EXPENSES', 'NET INCOME', 'EXCESS OR DEFICIT'],
            'TotalAssetsEOYAmt': ['TOTAL ASSETS', 'ASSETS TOTAL'],
            'TotalLiabilitiesEOYAmt': ['TOTAL LIABILITIES', 'LIABILITIES TOTAL'],
            'NetAssetsOrFundBalancesEOYAmt': ['NET ASSETS OR FUND BALANCES', 'TOTAL NET ASSETS', 'FUND BALANCES'],
            'TotalEmployeeCnt': ['TOTAL NUMBER OF EMPLOYEES', 'NUMBER OF EMPLOYEES', 'EMPLOYEES'],
            'TotalVolunteersCnt': ['TOTAL NUMBER OF VOLUNTEERS', 'NUMBER OF VOLUNTEERS', 'VOLUNTEERS'],
            'InformationTechnologyGrp': ['INFORMATION TECHNOLOGY', 'IT EXPENSES', 'TECHNOLOGY EXPENSE'],
            'OccupancyGrp': ['OCCUPANCY', 'RENT', 'OCCUPANCY EXPENSES'],
            'TravelGrp': ['TRAVEL', 'TRAVEL EXPENSES', 'TRAVEL COSTS']
        }
        
        # Add balance sheet patterns
        balance_sheet_patterns = {
            'CashNonInterestBearingEOY': ['CASH NON-INTEREST BEARING', 'CASH - NON-INTEREST BEARING', 'CASH END OF YEAR'],
            'AccountsReceivableEOY': ['ACCOUNTS RECEIVABLE', 'RECEIVABLES'],
            'AccountsPayableEOY': ['ACCOUNTS PAYABLE', 'ACCOUNTS PAYABLE AND ACCRUED EXPENSES', 'PAYABLES']
        }

        # Donor restriction patterns
        donor_restriction_patterns = {
            'WithoutDonorRestrictions': ['NO DONOR RESTRICTION', 'UNRESTRICTED NET ASSETS', 'WITHOUT DONOR RESTRICTIONS', 'NET ASSETS WITHOUT DONOR RESTRICTIONS'],
            'WithDonorRestrictions': ['DONOR RESTRICTION', 'PERMANENTLY RESTRICTED', 'TEMPORARILY RESTRICTED', 'WITH DONOR RESTRICTIONS', 'NET ASSETS WITH DONOR RESTRICTIONS']
        }

        # Process regular financial metrics
        for field, patterns in field_patterns.items():
            for pattern in patterns:
                found = False
                for i, line in enumerate(lines):
                    if pattern in line.upper():
                        # Look in current and next few lines for a value
                        for j in range(i, min(i + 5, len(lines))):
                            value = self._extract_numeric_value(lines[j])
                            if value:
                                metrics[field] = value
                                found = True
                                break
                    if found:
                        break

        # Process balance sheet items with special handling for EOY values
        for field, patterns in balance_sheet_patterns.items():
            for pattern in patterns:
                for i, line in enumerate(lines):
                    if pattern in line.upper():
                        # First try to find "End of Year" or "EOY" on the same line
                        if "END OF YEAR" in line.upper() or "EOY" in line.upper():
                            value = self._extract_numeric_value(line)
                            if value:
                                metrics[field] = value
                                break
                                
                        # If not on the same line, look for columns - try to find value in right-most position
                        else:
                            # Look for numeric values and take the right-most one (assuming it's EOY)
                            # This is based on the common format where BOY is left column, EOY is right column
                            values = self._extract_all_numeric_values(line)
                            if values and len(values) > 1:
                                metrics[field] = values[-1]  # Take the last (right-most) value
                                break
                            
                            # If still not found, check next line
                            if i + 1 < len(lines):
                                next_line = lines[i + 1]
                                values = self._extract_all_numeric_values(next_line)
                                if values and len(values) > 1:
                                    metrics[field] = values[-1]
                                    break

        # Process donor restrictions
        for field, patterns in donor_restriction_patterns.items():
            for pattern in patterns:
                for i, line in enumerate(lines):
                    if pattern in line.upper():
                        # Check if EOY/End of Year is in the line
                        if "END OF YEAR" in line.upper() or "EOY" in line.upper():
                            value = self._extract_numeric_value(line)
                            if value:
                                metrics[field] = value
                                break
                        
                        # Otherwise look in nearby lines
                        else:
                            # Look for line with END OF YEAR or EOY
                            for j in range(max(0, i - 3), min(i + 4, len(lines))):
                                if "END OF YEAR" in lines[j].upper() or "EOY" in lines[j].upper():
                                    value = self._extract_numeric_value(lines[j])
                                    if value:
                                        metrics[field] = value
                                        break
                                    
                            # If still not found, look for a line with numbers below the match
                            if field not in metrics:
                                for j in range(i + 1, min(i + 4, len(lines))):
                                    value = self._extract_numeric_value(lines[j])
                                    if value:
                                        metrics[field] = value
                                        break

        # Find total functional expenses with more comprehensive search
        for i, line in enumerate(lines):
            if 'TOTAL FUNCTIONAL EXPENSES' in line.upper() or 'STATEMENT OF FUNCTIONAL EXPENSES' in line.upper():
                # Search more extensively for management and fundraising amounts
                for j in range(i, min(i + 30, len(lines))):
                    current_line = lines[j].upper()
                    
                    # Management and general
                    if 'MANAGEMENT AND GENERAL' in current_line or 'MANAGEMENT & GENERAL' in current_line:
                        # Try current line first
                        value = self._extract_numeric_value(current_line)
                        if value:
                            metrics['ManagementAndGeneralAmt'] = value
                        else:
                            # Look at next line if current line doesn't have a value
                            if j + 1 < len(lines):
                                value = self._extract_numeric_value(lines[j + 1])
                                if value:
                                    metrics['ManagementAndGeneralAmt'] = value
                    
                    # Fundraising
                    if 'FUNDRAISING' in current_line and 'TOTAL' not in current_line:
                        # Try current line first
                        value = self._extract_numeric_value(current_line)
                        if value:
                            metrics['CYTotalFundraisingExpenseAmt'] = value
                        else:
                            # Look at next line if current line doesn't have a value
                            if j + 1 < len(lines):
                                value = self._extract_numeric_value(lines[j + 1])
                                if value:
                                    metrics['CYTotalFundraisingExpenseAmt'] = value
        
        return metrics

    def extract_executive_compensation(self, content, format_type):
        """Extract executive compensation data"""
        try:
            if format_type == 'xml':
                return self._extract_executive_compensation_xml(content)
            else:
                return self._extract_executive_compensation_txt(content)
        except Exception as e:
            logger.error(f"Error extracting executive compensation: {str(e)}")
            return []

    def _extract_executive_compensation_xml(self, tree):
        """Extract executive compensation from XML format"""
        root = tree.getroot()
        executives = []
        
        # Look for compensation data in Form 990 Part VII
        for person in root.findall('.//irs:Form990PartVIISectionAGrp', self.ns):
            name = person.find('.//irs:PersonNm', self.ns)
            title = person.find('.//irs:TitleTxt', self.ns)
            compensation = person.find('.//irs:ReportableCompFromOrgAmt', self.ns)
            
            if all(elem is not None for elem in [name, title, compensation]):
                if self._is_leadership_title(title.text):
                    executives.append({
                        'name': name.text,
                        'title': title.text,
                        'compensation': compensation.text
                    })
        
        return executives
        
    def _extract_executive_compensation_txt(self, content):
        """Extract executive compensation from TXT format"""
        executives = []
        lines = content.split('\n')
            
        current_person = {}
        for i, line in enumerate(lines):
            line = line.upper()
            # Look for sections that typically contain compensation information
            if 'FORM 990, PART VII' in line or 'COMPENSATION OF OFFICERS' in line:
                # Look through next several lines for compensation information
                for j in range(i, min(i + 100, len(lines))):
                    line = lines[j].strip().upper()
                    
                    # Check for leadership titles
                    if any(title in line for title in self.leadership_titles):
                        # Try to extract name, title, and compensation
                        parts = line.split()
                        # Look for dollar amounts
                        for k, part in enumerate(parts):
                            if '$' in part or (part.replace(',', '').isdigit() and len(part) > 4):
                                try:
                                    compensation = part.replace('$', '').replace(',', '')
                                    # Assume title is before compensation and name is at start
                                    title = ' '.join(parts[1:k])  # Skip first word (assume it's part of name)
                                    name = parts[0]  # Just take first word as name for simplicity
                                    
                                    executives.append({
                                        'name': name,
                                        'title': title,
                                        'compensation': compensation
                                    })
                                    break
                                except ValueError:
                                    continue
            
        return executives

    def extract_endowment_data(self, content, format_type):
        """Extract endowment data"""
        try:
            if format_type == 'xml':
                return self._extract_endowment_data_xml(content)
            else:
                return self._extract_endowment_data_txt(content)
        except Exception as e:
            logger.error(f"Error extracting endowment data: {str(e)}")
            return {}

    def _extract_endowment_data_xml(self, tree):
        """Extract endowment data from XML format"""
        root = tree.getroot()
        endowment_data = {}
        
        # Define field mapping
        field_mapping = {
            'BeginningYearBalanceAmt': 'BeginningYearBalanceAmt',
            'ContributionsAmt': 'ContributionsAmt',
            'InvestmentEarningsOrLossesAmt': 'InvestmentEarningsOrLossesAmt',
            'GrantsOrScholarshipsAmt': 'GrantsOrScholarshipsAmt',
            'OtherExpendituresAmt': 'OtherExpendituresAmt',
            'AdministrativeExpensesAmt': 'AdministrativeExpensesAmt',
            'EndYearBalanceAmt': 'EndYearBalanceAmt'
        }
        
        # Define year groups with their XML tags
        year_groups = [
            ('CYEndwmtFundGrp', 'Year_0'),
            ('CYMinus1YrEndwmtFundGrp', 'Year_1'),
            ('CYMinus2YrEndwmtFundGrp', 'Year_2'),
            ('CYMinus3YrEndwmtFundGrp', 'Year_3'),
            ('CYMinus4YrEndwmtFundGrp', 'Year_4')
        ]
        
        # First try to find Schedule D
        schedule_d = root.find('.//irs:IRS990ScheduleD', self.ns)
        if schedule_d is not None:
            root_to_search = schedule_d
        else:
            # If Schedule D is not found, search in the entire document
            root_to_search = root
            
        for group_tag, year_key in year_groups:
            year_data = {}
            # Search for the group in the current root
            group = root_to_search.find(f'.//irs:{group_tag}', self.ns)
            
            if group is not None:
                for field, xml_tag in field_mapping.items():
                    value = group.find(f'.//irs:{xml_tag}', self.ns)
                    if value is not None and value.text:
                        try:
                            # Convert to float to handle negative numbers properly
                            year_data[field] = str(float(value.text))
                        except ValueError:
                            year_data[field] = value.text
                    else:
                        year_data[field] = None
                        
                if any(year_data.values()):  # Only add if we found any data
                    endowment_data[year_key] = year_data
        

        
        return endowment_data
    
    def _extract_all_numeric_values(self, line):
        """Extract all numeric values from a line of text"""
        if not line:
            return []
        
        values = []
        # Clean the line first
        clean_line = line.replace('$', '').replace(',', '')
        words = clean_line.split()
        
        for word in words:
            try:
                # Try to convert to float
                value = float(''.join(c for c in word if c.isdigit() or c in '.-'))
                if value != 0:  # Skip zero values as they're often not meaningful
                    values.append(str(value))
            except ValueError:
                continue
        
        return values
        
    def _extract_numeric_value(self, line):
        """Extract numeric value from text line - improved version"""
        if not line:
            return None
    
        # Clean the line
        clean_line = line.replace('$', '').replace(',', '')
        
        # Try a few common patterns for financial data in forms
        
        # Pattern 1: Last number on the line is often the value
        words = clean_line.split()
        for word in reversed(words):
            try:
                # Extract only digits and decimal points
                numeric_part = ''.join(c for c in word if c.isdigit() or c in '.-')
                if numeric_part:
                    return str(float(numeric_part))
            except ValueError:
                continue
        
        # Pattern 2: Look for parentheses which often indicate negative numbers
        import re
        pattern = r'\(([0-9,]+(?:\.[0-9]+)?)\)'
        matches = re.findall(pattern, line)
        if matches:
            try:
                # Negative value in parentheses
                return str(-float(matches[0].replace(',', '')))
            except ValueError:
                pass
        
        # Pattern 3: Look for dollar sign followed by number
        pattern = r'\$\s*([0-9,]+(?:\.[0-9]+)?)'
        matches = re.findall(pattern, line)
        if matches:
            try:
                return str(float(matches[0].replace(',', '')))
            except ValueError:
                pass
        
        return None


    def extract_endowment_data_txt(self, content):
        """Extract endowment data from TXT format"""
        endowment_data = {}
        lines = content.split('\n')
        
        # Look for endowment section
        in_endowment_section = False
        current_year_data = {}
        
        for i, line in enumerate(lines):
            line_upper = line.upper()
            
            # Check for start of endowment section
            if 'ENDOWMENT FUNDS' in line_upper or 'SCHEDULE D, PART V' in line_upper:
                in_endowment_section = True
                continue
                
            if in_endowment_section:
                # Check for end of section
                if 'PART VI' in line_upper or 'STATEMENT OF REVENUE' in line_upper:
                    in_endowment_section = False
                    continue
                    
                # Look for specific endowment data fields
                field_patterns = {
                    'BeginningYearBalanceAmt': ['BEGINNING OF YEAR', 'BEGINNING BALANCE'],
                    'ContributionsAmt': ['CONTRIBUTIONS', 'ADDITIONS'],
                    'InvestmentEarningsOrLossesAmt': ['INVESTMENT EARNINGS', 'NET INVESTMENT EARNINGS', 'INVESTMENT GAINS'],
                    'GrantsOrScholarshipsAmt': ['GRANTS', 'SCHOLARSHIPS', 'GRANTS OR SCHOLARSHIPS'],
                    'OtherExpendituresAmt': ['OTHER EXPENDITURES', 'OTHER EXPENSES'],
                    'AdministrativeExpensesAmt': ['ADMINISTRATIVE', 'ADMIN EXPENSES'],
                    'EndYearBalanceAmt': ['END OF YEAR', 'ENDING BALANCE']
                }
                
                for field, patterns in field_patterns.items():
                    if any(pattern in line_upper for pattern in patterns):
                        # Look for numeric values
                        value = self._extract_numeric_value(line)
                        if value:
                            current_year_data[field] = value
        
        # If we found any endowment data, add it
        if current_year_data:
            endowment_data['Year_0'] = current_year_data
        
        return endowment_data

    def _is_leadership_title(self, title):
        """Check if a title matches leadership positions"""
        if not title:
            return False
        
        return any(leadership_title in title.upper() for leadership_title in self.leadership_titles)

    def _extract_numeric_value(self, line):
        """Extract numeric value from text line"""
        try:
            # Remove common currency formatting
            clean_line = line.replace('$', '').replace(',', '').strip()
            # Find last sequence of digits (possibly with decimal point)
            words = clean_line.split()
            for word in reversed(words):
                try:
                    return str(float(word))
                except ValueError:
                    continue
            return None
        except Exception:
            return None



class ExcelOutputHandler:
    """Handles formatting and writing data to Excel in vertical format with metrics as rows"""

    def __init__(self, output_path):
        self.output_path = output_path
        self.field_mapping = {
            # Regular financial fields
            'Total Revenue': 'CYTotalRevenueAmt',
            'Total Contributions': 'CYContributionsGrantsAmt',
            'Investment Income': 'CYInvestmentIncomeAmt',
            'Grants and Salaries': 'CYGrantsAndSimilarPaidAmt',
            'Salaries Other': 'CYSalariesCompEmpBnftPaidAmt',
            'Total Expenses': 'CYTotalExpensesAmt',
            'Program Service Expenses': 'TotalProgramServiceExpensesAmt',
            'Management': 'ManagementAndGeneralAmt',
            'Fundraising': 'CYTotalFundraisingExpenseAmt',
            'Revenue Less': 'CYRevenuesLessExpensesAmt',
            'Information Technology': 'InformationTechnologyGrp',
            'Accounting': 'FeesForServicesAccountingGrp',
            'Occupancy': 'OccupancyGrp',
            'Travel': 'TravelGrp',
            'Number of Employees': 'TotalEmployeeCnt',
            'Number of Volunteers': 'TotalVolunteersCnt',
            'Other Revenue': 'CYOtherRevenueAmt',
            'Program Service': 'CYProgramServiceRevenueAmt',
            'Other Expenses': 'CYOtherExpensesAmt',
            'Cash NonInterest Bearing': 'CashNonInterestBearingEOY',
            'Accounts Receivable': 'AccountsReceivableEOY',
            'Accounts Payable': 'AccountsPayableEOY',
            'Total Assets': 'TotalAssetsEOYAmt',
            'Total Liabilities': 'TotalLiabilitiesEOYAmt',
            'Net Assets': 'NetAssetsOrFundBalancesEOYAmt',
            'Net Assets Without Donor Restrictions': 'WithoutDonorRestrictions',
            'Net Assets With Donor Restrictions': 'WithDonorRestrictions',
            # Endowment fields
            'Endowment Beginning Balance': 'BeginningYearBalanceAmt',
            'Endowment Contributions': 'ContributionsAmt',
            'Endowment Investment Earnings': 'InvestmentEarningsOrLossesAmt',
            'Endowment Grants': 'GrantsOrScholarshipsAmt',
            'Endowment Other Expenditures': 'OtherExpendituresAmt',
            'Endowment Admin Expenses': 'AdministrativeExpensesAmt',
            'Endowment Ending Balance': 'EndYearBalanceAmt',
        }

    def clean_sheet_name(self, name):
        """Clean sheet name to comply with Excel's 31-character limit and other restrictions"""
        if not name:
            return "Sheet"

        # Remove invalid characters for Excel sheet names
        invalid_chars = r'[\\/*?:[\]]'
        name = re.sub(invalid_chars, '', name)
        
        # Remove leading/trailing spaces and collapse multiple spaces
        name = ' '.join(name.split())
        
        # If name is still too long, intelligently truncate it
        if len(name) > 31:
            # Try to find a word boundary to break at
            words = name.split()
            shortened_name = ""
            for word in words:
                if len(shortened_name + " " + word) > 28:  # Leave room for ellipsis
                    break
                shortened_name += (" " + word if shortened_name else word)
            
            name = shortened_name.strip() + "..."
        
        # Final verification of length
        if len(name) > 31:
            name = name[:28] + "..."
        
        # Ensure name is not empty and doesn't start/end with spaces
        name = name.strip()
        if not name:
            name = "Sheet"
            
        return name

    def write_to_excel(self, org_dfs):
        """Write data to Excel with each organization in its own sheet"""
        try:
            if not org_dfs:
                logger.error("No data to write to Excel")
                return False
                
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Track used sheet names to avoid duplicates
                used_names = set()
                
                for org_name, org_info in org_dfs.items():
                    df = org_info['data']
                    ntee_category = org_info['ntee_category']
                    
                    # Get clean sheet name
                    base_name = self.clean_sheet_name(org_name)
                    sheet_name = base_name
                    
                    # Handle duplicate sheet names
                    counter = 1
                    while sheet_name.lower() in used_names:
                        # If we need to add a number, make sure we have room
                        base_truncated = base_name[:27] if len(base_name) > 27 else base_name
                        sheet_name = f"{base_truncated}_{counter}"
                        counter += 1
                    
                    used_names.add(sheet_name.lower())
                    
                    # Create the worksheet
                    worksheet = writer.book.create_sheet(sheet_name)
                    worksheet['A1'] = f'NTEE Category: {ntee_category}'
                    df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
                    
                    # Format the worksheet
                    self._format_worksheet(worksheet, df)
                
                # Remove default sheet if it exists
                if 'Sheet' in writer.book.sheetnames:
                    del writer.book['Sheet']
            
            return True
            
        except Exception as e:
            logger.error(f"Error writing to Excel: {str(e)}")
            raise
    
    def format_value(self, value, metric_name):
        """
        Format numeric values appropriately
        Args:
            value: The value to format
            metric_name: The name of the metric (to determine formatting)
        """
        if value is None or value == 'Not found':
            return None
            
        # List of metrics that should not be in dollars
        non_dollar_metrics = [
            'Number of Employees',
            'Number of Volunteers'
        ]
        
        if isinstance(value, str):
            try:
                # Remove any existing formatting
                clean_value = ''.join(c for c in value if c.isdigit() or c in '.-')
                numeric_value = float(clean_value)
                
                # Return raw numbers for employee/volunteer counts
                if metric_name in non_dollar_metrics:
                    return numeric_value
                else:
                    # Return dollar values for everything else
                    return numeric_value
                    
            except (ValueError, TypeError):
                return value
                
        return value

    def consolidate_data(self, org_data):
        """Consolidate data into vertical format with metrics as rows"""
        org_dfs = {}
        years_range = list(range(2018, 2023))  # 2018-2022
        
        # Create a mapping dictionary to track all variations of an org name
        name_mapping = {}
        
        # First pass: build name mapping
        for ntee_category, orgs in org_data.items():
            for org_name in orgs.keys():
                normalized_name = self._normalize_org_name(org_name)
                name_mapping[org_name] = normalized_name
        
        # Second pass: consolidate data using normalized names
        normalized_org_data = {}
        for ntee_category, orgs in org_data.items():
            for org_name, years_data in orgs.items():
                normalized_name = name_mapping[org_name]
                
                # Create category if needed
                if ntee_category not in normalized_org_data:
                    normalized_org_data[ntee_category] = {}
                
                # Create organization if needed
                if normalized_name not in normalized_org_data[ntee_category]:
                    normalized_org_data[ntee_category][normalized_name] = []
                
                # Add the data
                normalized_org_data[ntee_category][normalized_name].extend(years_data)
        
        # Now process using normalized data structure
        for ntee_category, orgs in normalized_org_data.items():
            for normalized_name, years_data in orgs.items():
                # Use first non-normalized name for display
                display_name = next((name for name, norm in name_mapping.items() 
                                if norm == normalized_name), normalized_name)
                
                # Initialize dictionary to store metrics by year
                metrics_by_year = {year: {} for year in years_range}
                
                # Process each year's data
                for year_data in years_data:
                    try:
                        tax_year = year_data.get('tax_year', 'Unknown')
                        if tax_year != 'Unknown':
                            tax_year = int(tax_year)
                            if tax_year in years_range:
                                metrics = year_data.get('financial_metrics', {})
                                endowment_data = year_data.get('endowment_data', {})
                                
                                # Store metrics for this year
                                year_metrics = {}
                                
                                # Process metrics in the order of field_mapping
                                for display_col, field_name in self.field_mapping.items():
                                    if display_col.startswith('Endowment '):
                                        # Handle endowment data
                                        if endowment_data and 'Year_0' in endowment_data:
                                            current_year_endowment = endowment_data['Year_0']
                                            value = current_year_endowment.get(field_name, None)
                                            if value is not None:
                                                year_metrics[display_col] = self.format_value(value, display_col)
                                    else:
                                        # Handle regular financial metrics
                                        value = metrics.get(field_name, None)
                                        year_metrics[display_col] = self.format_value(value, display_col)
                                
                                metrics_by_year[tax_year] = year_metrics
                    except Exception as e:
                        logger.error(f"Error processing year data for {display_name}: {str(e)}")
                        continue
                
                # Create rows for DataFrame using field_mapping order
                rows = []
                for metric_display_name in self.field_mapping.keys():
                    row = {'Metric': metric_display_name}
                    for year in years_range:
                        row[str(year)] = metrics_by_year[year].get(metric_display_name, None)
                    rows.append(row)
                
                # Create DataFrame if we have rows
                if rows:
                    df = pd.DataFrame(rows)
                    df.insert(0, 'Organization', display_name)
                    
                    # Store DataFrame with NTEE category
                    org_dfs[display_name] = {
                        'data': df,
                        'ntee_category': ntee_category
                    }
        
        return org_dfs
        

    def _normalize_org_name(self, name):
        """Normalize organization name for consistent matching"""
        if not name:
            return "Unknown"
            
        # Strip extra whitespace and convert to uppercase
        normalized = ' '.join(name.upper().split())
        
        # Remove common suffixes
        common_suffixes = [' INC', ' LLC', ' FOUNDATION', ' CORP', ' CORPORATION', ' LTD', ' INCORPORATED']
        for suffix in common_suffixes:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)]
        
        # Handle special case for "UNIVERSITY"
        if "UNIVERSITY" in normalized:
            # Extract the main university name
            parts = normalized.split()
            if len(parts) > 1 and parts[-1] == "UNIVERSITY":
                normalized = ' '.join(parts[:-1]) + " UNIVERSITY"
        
        return normalized

    def write_to_excel(self, org_dfs):
        """Write data to Excel with each organization in its own sheet"""
        try:
            if not org_dfs:
                logger.error("No data to write to Excel")
                return False
                
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Create first sheet before removing default
                first_org = next(iter(org_dfs.items()))
                first_org_name, first_org_info = first_org
                first_df = first_org_info['data']
                first_ntee = first_org_info['ntee_category']
                first_sheet_name = self.clean_sheet_name(first_org_name)
                
                # Create and format first sheet
                worksheet = writer.book.create_sheet(first_sheet_name)
                worksheet['A1'] = f'NTEE Category: {first_ntee}'
                first_df.to_excel(writer, sheet_name=first_sheet_name, startrow=2, index=False)
                
                # Format first worksheet
                worksheet = writer.book[first_sheet_name]
                self._format_worksheet(worksheet, first_df)
                
                # Now safe to remove default sheet
                if 'Sheet' in writer.book.sheetnames:
                    std = writer.book['Sheet']
                    writer.book.remove(std)
                
                # Process remaining organizations
                remaining_orgs = list(org_dfs.items())[1:]
                for org_name, org_info in remaining_orgs:
                    df = org_info['data']
                    ntee_category = org_info['ntee_category']
                    sheet_name = self.clean_sheet_name(org_name)
                    
                    # Write the sheet
                    worksheet = writer.book.create_sheet(sheet_name)
                    worksheet['A1'] = f'NTEE Category: {ntee_category}'
                    df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
                    
                    # Format the worksheet
                    self._format_worksheet(worksheet, df)
            
            return True
            
        except Exception as e:
            logger.error(f"Error writing to Excel: {str(e)}")
            raise


    def _normalize_org_name(self, name):
        """Normalize organization name for consistent matching"""
        if not name:
            return "Unknown"
            
        normalized = name.upper()
        common_suffixes = [' INC', ' LLC', ' FOUNDATION', ' CORP', ' CORPORATION', ' LTD', ' INCORPORATED']
        for suffix in common_suffixes:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)]
        
        return ' '.join(normalized.split())

    def _format_worksheet(self, worksheet, df):
        """Helper method to format worksheet"""
        # List of metrics that should not be in dollars
        non_dollar_metrics = [
            'Number of Employees',
            'Number of Volunteers'
        ]
        
        # Format columns
        for idx, col in enumerate(df.columns):
            # Set column width
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx//26) + chr(65 + (idx % 26))
            worksheet.column_dimensions[col_letter].width = max_length + 2
            
            # Format numeric columns (year columns)
            if col.isdigit():
                for row in range(4, len(df) + 4):  # Skip header and NTEE category
                    cell = worksheet.cell(row=row, column=idx + 1)
                    metric_name = worksheet.cell(row=row, column=1).value  # Get metric name from first column
                    
                    try:
                        if pd.notna(cell.value):
                            value = float(cell.value)
                            if metric_name in non_dollar_metrics:
                                cell.value = round(value)  # Force integer for counts
                                cell.number_format = '#,##0'  # Regular number format for counts
                            else:
                                cell.value = round(value)  # Force integer for dollar amounts
                                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Accounting format
                    except (ValueError, TypeError):
                        continue
        
        # Format NTEE category row
        from openpyxl.styles import Font
        worksheet['A1'].font = Font(bold=True)


def main():
    # Initialize components
    parser = NonprofitParser()
    extractor = FinancialDataExtractor()
    excel_handler = ExcelOutputHandler(r'C:\Users\aronc\OneDrive\Documents\PushExcel.xlsx')
    
    # Dictionary to store all org data by NTEE category
    all_org_data = {}
    
    while True:
        # Get organization URL from user
        org_url = input("\nEnter the ProPublica organization URL (or type 'done' to finish): ")
        
        if org_url.lower() == 'done':
            break
            
        try:
            # Get NTEE category and XML links
            ntee_category, xml_links = parser.scraper.get_organization_links(org_url)
            
            # Initialize category if needed
            if ntee_category not in all_org_data:
                all_org_data[ntee_category] = {}
            
            logger.info(f"\nProcessing organization in category: {ntee_category}")
            
            for url in xml_links:
                try:
                    # Parse basic information
                    result = parser.process_url(url)
                    org_name = result['organization_name']
                    
                    # Initialize organization in data structure if needed
                    if org_name not in all_org_data[ntee_category]:
                        all_org_data[ntee_category][org_name] = []
                    
                    # Extract financial data
                    result['financial_metrics'] = extractor.extract_financial_metrics(
                        result['parsed_content'],
                        result['format']
                    )
                    
                    # Extract executive compensation
                    result['executive_compensation'] = extractor.extract_executive_compensation(
                        result['parsed_content'],
                        result['format']
                    )
                    
                    
                    all_org_data[ntee_category][org_name].append(result)
                    logger.info(f"Successfully processed {url}")
                    
                except Exception as e:
                    logger.error(f"Error processing {url}: {str(e)}")
                    continue
                    
        except Exception as e:
            logger.error(f"Error processing organization: {str(e)}")
            continue
    
    # After all URLs are processed, write to Excel
    if all_org_data:
        category_dfs = excel_handler.consolidate_data(all_org_data)
        if excel_handler.write_to_excel(category_dfs):
            logger.info(f"\nSuccessfully wrote data to {excel_handler.output_path}")
        else:
            logger.error("Failed to write to Excel file")
    else:
        logger.error("No data was processed successfully")



if __name__ == "__main__":
    main()
