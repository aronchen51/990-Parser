import xml.etree.ElementTree as ET
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime
from collections import defaultdict
import logging
from bs4 import BeautifulSoup
import re
import os
import openpyxl


# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ProPublicaScraper:
    """Scraper for ProPublica nonprofit search pages"""
    
    def __init__(self):
        self.base_url = "https://projects.propublica.org"
    
    def get_organization_links(self, main_url):
        """
        Extract XML download links and NTEE category from organization's main page
        Returns: tuple (ntee_category, list_of_xml_urls)
        """
        try:
            # Fetch the main page
            response = requests.get(main_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract NTEE category
            ntee_elem = soup.find('p', class_='ntee-category')
            if ntee_elem:
                ntee_category = ntee_elem.text.split(':', 1)[1].strip().split('/')[0].strip()
            else:
                ntee_category = "Unknown"
            
            # Find all XML download links
            xml_links = []
            for link in soup.find_all('a', class_='btn', target='_blank'):
                if 'XML' in link.text:
                    object_id = link['href'].split('object_id=')[1]
                    full_url = f"{self.base_url}/nonprofits/download-xml?object_id={object_id}"
                    xml_links.append(full_url)
            
            # Sort links by object_id (which contains year) and take most recent 5
            xml_links.sort(reverse=True)
            xml_links = xml_links[:5]
            
            return ntee_category, xml_links
            
        except requests.RequestException as e:
            logger.error(f"Error fetching organization page: {str(e)}")
            raise


class NonprofitPFParser:
    """Parser for 990-PF private foundation financial data from ProPublica URLs"""
    
    def __init__(self):
        self.scraper = ProPublicaScraper()
        self.ns = {'irs': 'http://www.irs.gov/efile'}

    def fetch_content(self, url):
        """Fetch content from URL with error handling"""
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response.content
        except requests.RequestException as e:
            logger.error(f"Failed to fetch from URL {url}: {str(e)}")
            raise

    def detect_format(self, content):
        """
        Detect whether content is XML or TXT format
        Returns tuple of (format_type, parsed_content)
        """
        try:
            # First try parsing as XML
            tree = ET.parse(BytesIO(content))
            logger.info("Successfully parsed as XML")
            return 'xml', tree
        except ET.ParseError:
            # If XML parsing fails, try TXT format
            try:
                text_content = content.decode('utf-8', errors='ignore')
                # Check for common TXT format markers
                if any(marker in text_content.upper() for marker in 
                      ['RETURN HEADER', 'FORM 990-PF', 'EIN:']):
                    logger.info("Successfully parsed as TXT")
                    return 'txt', text_content
                else:
                    raise ValueError("Content doesn't match expected formats")
            except Exception as e:
                logger.error(f"Format detection failed: {str(e)}")
                raise

    def get_tax_year(self, content, format_type):
        """Extract tax year from content and adjust it to reflect reporting year"""
        try:
            if format_type == 'xml':
                root = content.getroot()
                # Try multiple possible locations for tax year
                tax_period = root.find('.//irs:TaxPeriodEndDt', self.ns)
                if tax_period is not None and tax_period.text:
                    # Subtract 1 from the tax year to get the reporting year
                    return str(int(datetime.strptime(tax_period.text, '%Y-%m-%d').year) - 1)
                
                tax_year = root.find('.//irs:TaxYr', self.ns)
                if tax_year is not None and tax_year.text:
                    # Subtract 1 from the tax year to get the reporting year
                    return str(int(tax_year.text) - 1)
            else:
                # Search for year in TXT content
                text_lines = content.split('\n')
                for line in text_lines:
                    if 'Tax Period Begin' in line:
                        # Extract first 4-digit number found and subtract 1
                        for word in line.split():
                            if word.isdigit() and len(word) == 4:
                                return str(int(word) - 1)
            
            logger.warning("Could not find tax year")
            return "Unknown"
        except Exception as e:
            logger.error(f"Error extracting tax year: {str(e)}")
            return "Unknown"

    def get_organization_name(self, content, format_type):
        """Extract and normalize organization name from content"""
        try:
            raw_name = None
            if format_type == 'xml':
                root = content.getroot()
                # Try multiple possible locations for organization name
                for path in [
                    './/irs:BusinessName/irs:BusinessNameLine1Txt',
                    './/irs:ReturnHeader/irs:Filer/irs:BusinessName/irs:BusinessNameLine1Txt'
                ]:
                    name = root.find(path, self.ns)
                    if name is not None and name.text:
                        raw_name = name.text
                        break
            else:
                # Search for organization name in TXT content
                text_lines = content.split('\n')
                for line in text_lines:
                    if 'Name of Organization:' in line or 'NAME OF ORGANIZATION:' in line:
                        raw_name = line.split(':', 1)[1].strip()
                        break
            
            if not raw_name:
                logger.warning("Could not find organization name")
                return "Unknown Organization"
                
            return raw_name
            
        except Exception as e:
            logger.error(f"Error extracting organization name: {str(e)}")
            return "Unknown Organization"
    
    def process_url(self, url):
        """Process a single URL and return basic information"""
        try:
            content = self.fetch_content(url)
            format_type, parsed_content = self.detect_format(content)
            
            # Initialize financial data extractor
            extractor = FinancialDataExtractorPF()
            
            result = {
                'url': url,
                'format': format_type,
                'tax_year': self.get_tax_year(parsed_content, format_type),
                'organization_name': self.get_organization_name(parsed_content, format_type),
                'parsed_content': parsed_content,
            }
            
            # Extract financial metrics
            result['financial_metrics'] = extractor.extract_financial_metrics(
                parsed_content,
                format_type
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing URL {url}: {str(e)}")
            raise


class FinancialDataExtractorPF:
    """Extracts financial data from parsed 990-PF private foundation documents"""
    
    def __init__(self):
        self.ns = {'irs': 'http://www.irs.gov/efile'}
        
    def extract_financial_metrics(self, content, format_type):
        """Extract basic financial metrics"""
        try:
            if format_type == 'xml':
                return self._extract_financial_metrics_xml(content)
            else:
                return self._extract_financial_metrics_txt(content)
        except Exception as e:
            logger.error(f"Error extracting financial metrics: {str(e)}")
            return {}

    def _extract_financial_metrics_xml(self, tree):
        """Extract financial metrics from 990-PF XML format"""
        root = tree.getroot()
        metrics = {}

        # 990-PF specific field mappings
        pf_elements = {
            # Revenue elements (from AnalysisOfRevenueAndExpenses)
            'TotalRevAndExpnssAmt': 'Total Revenue',
            'ContriRcvdRevAndExpnssAmt': 'Total Contributions',
            'InterestOnSavRevAndExpnssAmt': 'Interest Income',
            'DividendsRevAndExpnssAmt': 'Dividend Income',
            'NetGainSaleAstRevAndExpnssAmt': 'Net Capital Gains',
            'OtherIncomeRevAndExpnssAmt': 'Other Revenue',
            
            # Expense elements
            'TotalExpensesRevAndExpnssAmt': 'Total Expenses',
            'ContriPaidRevAndExpnssAmt': 'Contributions Paid',
            'CompOfcrDirTrstRevAndExpnssAmt': 'Officer Compensation',
            'OthEmplSlrsWgsRevAndExpnssAmt': 'Other Employee Salaries',
            'AccountingFeesRevAndExpnssAmt': 'Accounting',
            'OccupancyRevAndExpnssAmt': 'Occupancy',
            'TravConfMeetingRevAndExpnssAmt': 'Travel',
            'OtherExpensesRevAndExpnssAmt': 'Other Expenses',
            'ExcessRevenueOverExpensesAmt': 'Revenue Less Expenses',
        }

        # Balance sheet elements (from Form990PFBalanceSheetsGrp)
        balance_sheet_elements = {
            'CashEOYAmt': 'Cash Noninterest Bearing',
            'TotalAssetsEOYAmt': 'Total Assets',
            'AccountsPayableEOYAmt': 'Accounts Payable',
            'TotalLiabilitiesEOYAmt': 'Total Liabilities',
            'TotNetAstOrFundBalancesEOYAmt': 'Net Assets',
            'NoDonorRstrNetAssestsEOYAmt': 'Net Assets Without Donor Restrictions',
            'DonorRstrNetAssetsEOYAmt': 'Net Assets With Donor Restrictions',
        }

        # Extract revenue and expense metrics from AnalysisOfRevenueAndExpenses
        analysis_section = root.find('.//irs:AnalysisOfRevenueAndExpenses', self.ns)
        if analysis_section is not None:
            for xml_element, display_name in pf_elements.items():
                element = analysis_section.find(f'.//irs:{xml_element}', self.ns)
                if element is not None and element.text:
                    metrics[display_name] = element.text
                else:
                    metrics[display_name] = 'Not found'

        # Extract balance sheet metrics
        balance_sheet = root.find('.//irs:Form990PFBalanceSheetsGrp', self.ns)
        if balance_sheet is not None:
            for xml_element, display_name in balance_sheet_elements.items():
                element = balance_sheet.find(f'.//irs:{xml_element}', self.ns)
                if element is not None and element.text:
                    metrics[display_name] = element.text
                else:
                    metrics[display_name] = 'Not found'

        # Calculate combined investment income
        interest_income = metrics.get('Interest Income', '0')
        dividend_income = metrics.get('Dividend Income', '0')
        capital_gains = metrics.get('Net Capital Gains', '0')
        
        try:
            total_investment_income = (
                int(interest_income) if interest_income != 'Not found' else 0
            ) + (
                int(dividend_income) if dividend_income != 'Not found' else 0
            ) + (
                int(capital_gains) if capital_gains != 'Not found' else 0
            )
            metrics['Investment Income'] = str(total_investment_income)
        except (ValueError, TypeError):
            metrics['Investment Income'] = 'Not found'

        # Calculate combined grants and salaries
        contributions_paid = metrics.get('Contributions Paid', '0')
        officer_comp = metrics.get('Officer Compensation', '0')
        employee_salaries = metrics.get('Other Employee Salaries', '0')
        
        try:
            total_grants_salaries = (
                int(contributions_paid) if contributions_paid != 'Not found' else 0
            ) + (
                int(officer_comp) if officer_comp != 'Not found' else 0
            ) + (
                int(employee_salaries) if employee_salaries != 'Not found' else 0
            )
            metrics['Grants and Salaries'] = str(total_grants_salaries)
        except (ValueError, TypeError):
            metrics['Grants and Salaries'] = 'Not found'

        return metrics

    def _extract_financial_metrics_txt(self, content):
        """Extract financial metrics from 990-PF TXT format"""
        metrics = {}
        lines = content.split('\n')
        
        # 990-PF specific field patterns for TXT format
        field_patterns = {
            'Total Revenue': ['TOTAL REVENUE', 'REVENUE TOTAL'],
            'Total Contributions': ['CONTRIBUTIONS RECEIVED', 'GIFTS RECEIVED', 'CONTRIBUTIONS AND GIFTS'],
            'Investment Income': ['INVESTMENT INCOME', 'DIVIDENDS AND INTEREST', 'NET INVESTMENT INCOME'],
            'Interest Income': ['INTEREST ON SAVINGS', 'INTEREST INCOME'],
            'Dividend Income': ['DIVIDENDS', 'DIVIDEND INCOME'],
            'Net Capital Gains': ['NET GAIN FROM SALE', 'CAPITAL GAINS', 'NET GAINS'],
            'Other Revenue': ['OTHER INCOME'],
            'Total Expenses': ['TOTAL EXPENSES', 'TOTAL OPERATING EXPENSES'],
            'Contributions Paid': ['CONTRIBUTIONS PAID', 'GRANTS PAID'],
            'Officer Compensation': ['COMPENSATION OF OFFICERS', 'OFFICER COMPENSATION'],
            'Other Employee Salaries': ['OTHER EMPLOYEE SALARIES', 'SALARIES AND WAGES'],
            'Accounting': ['ACCOUNTING FEES'],
            'Occupancy': ['OCCUPANCY'],
            'Travel': ['TRAVEL', 'CONFERENCES MEETINGS'],
            'Other Expenses': ['OTHER EXPENSES'],
            'Revenue Less Expenses': ['EXCESS OF REVENUE', 'NET INCOME'],
            'Cash Noninterest Bearing': ['CASH NON-INTEREST', 'CASH END OF YEAR'],
            'Total Assets': ['TOTAL ASSETS'],
            'Accounts Payable': ['ACCOUNTS PAYABLE'],
            'Total Liabilities': ['TOTAL LIABILITIES'],
            'Net Assets': ['NET ASSETS', 'FUND BALANCES'],
            'Net Assets Without Donor Restrictions': ['NO DONOR RESTRICTION', 'UNRESTRICTED'],
            'Net Assets With Donor Restrictions': ['DONOR RESTRICTION', 'RESTRICTED'],
        }

        # Process each field pattern
        for field, patterns in field_patterns.items():
            for pattern in patterns:
                found = False
                for i, line in enumerate(lines):
                    if pattern in line.upper():
                        # Look in current and next few lines for a value
                        for j in range(i, min(i + 5, len(lines))):
                            value = self._extract_numeric_value(lines[j])
                            if value:
                                metrics[field] = value
                                found = True
                                break
                    if found:
                        break

        return metrics

    def _extract_numeric_value(self, line):
        """Extract numeric value from text line"""
        if not line:
            return None
    
        # Clean the line
        clean_line = line.replace('$', '').replace(',', '')
        
        # Try to find the last number on the line (common in financial forms)
        words = clean_line.split()
        for word in reversed(words):
            try:
                # Extract only digits and decimal points
                numeric_part = ''.join(c for c in word if c.isdigit() or c in '.-')
                if numeric_part:
                    return str(float(numeric_part))
            except ValueError:
                continue
        
        # Look for parentheses which often indicate negative numbers
        import re
        pattern = r'\(([0-9,]+(?:\.[0-9]+)?)\)'
        matches = re.findall(pattern, line)
        if matches:
            try:
                # Negative value in parentheses
                return str(-float(matches[0].replace(',', '')))
            except ValueError:
                pass
        
        return None


class ExcelOutputHandlerPF:
    """Handles formatting and writing 990-PF data to Excel in vertical format with metrics as rows"""

    def __init__(self, output_path):
        self.output_path = output_path
        self.field_mapping = {
            # Revenue fields
            'Total Revenue': 'Total Revenue',
            'Total Contributions': 'Total Contributions',
            'Investment Income': 'Investment Income',
            # Expense fields
            'Grants and Salaries': 'Grants and Salaries',
            'Total Expenses': 'Total Expenses',
            'Revenue Less Expenses': 'Revenue Less Expenses',
            # Operational categories we're keeping
            'Accounting': 'Accounting',
            'Occupancy': 'Occupancy',
            'Travel': 'Travel',
            # Other financial
            'Other Revenue': 'Other Revenue',
            'Other Expenses': 'Other Expenses',
            # Balance sheet
            'Cash Noninterest Bearing': 'Cash Noninterest Bearing',
            'Accounts Payable': 'Accounts Payable',
            'Total Assets': 'Total Assets',
            'Total Liabilities': 'Total Liabilities',
            'Net Assets': 'Net Assets',
            'Net Assets Without Donor Restrictions': 'Net Assets Without Donor Restrictions',
        }

    def clean_sheet_name(self, name):
        """Clean sheet name to comply with Excel's 31-character limit and other restrictions"""
        if not name:
            return "Sheet"

        # Remove invalid characters for Excel sheet names
        invalid_chars = r'[\\/*?:[\]]'
        name = re.sub(invalid_chars, '', name)
        
        # Remove leading/trailing spaces and collapse multiple spaces
        name = ' '.join(name.split())
        
        # If name is still too long, intelligently truncate it
        if len(name) > 31:
            # Try to find a word boundary to break at
            words = name.split()
            shortened_name = ""
            for word in words:
                if len(shortened_name + " " + word) > 28:  # Leave room for ellipsis
                    break
                shortened_name += (" " + word if shortened_name else word)
            
            name = shortened_name.strip() + "..."
        
        # Final verification of length
        if len(name) > 31:
            name = name[:28] + "..."
        
        # Ensure name is not empty and doesn't start/end with spaces
        name = name.strip()
        if not name:
            name = "Sheet"
            
        return name

    def format_value(self, value, metric_name):
        """
        Format numeric values appropriately
        Args:
            value: The value to format
            metric_name: The name of the metric (to determine formatting)
        """
        if value is None or value == 'Not found':
            return None
            
        if isinstance(value, str):
            try:
                # Remove any existing formatting
                clean_value = ''.join(c for c in value if c.isdigit() or c in '.-')
                numeric_value = float(clean_value)
                return numeric_value
                    
            except (ValueError, TypeError):
                return value
                
        return value

    def consolidate_data(self, org_data):
        """Consolidate data into vertical format with metrics as rows"""
        org_dfs = {}
        years_range = list(range(2018, 2024))  # 2018-2022
        
        # Create a mapping dictionary to track all variations of an org name
        name_mapping = {}
        
        # First pass: build name mapping
        for ntee_category, orgs in org_data.items():
            for org_name in orgs.keys():
                normalized_name = self._normalize_org_name(org_name)
                name_mapping[org_name] = normalized_name
        
        # Second pass: consolidate data using normalized names
        normalized_org_data = {}
        for ntee_category, orgs in org_data.items():
            for org_name, years_data in orgs.items():
                normalized_name = name_mapping[org_name]
                
                # Create category if needed
                if ntee_category not in normalized_org_data:
                    normalized_org_data[ntee_category] = {}
                
                # Create organization if needed
                if normalized_name not in normalized_org_data[ntee_category]:
                    normalized_org_data[ntee_category][normalized_name] = []
                
                # Add the data
                normalized_org_data[ntee_category][normalized_name].extend(years_data)
        
        # Now process using normalized data structure
        for ntee_category, orgs in normalized_org_data.items():
            for normalized_name, years_data in orgs.items():
                # Use first non-normalized name for display
                display_name = next((name for name, norm in name_mapping.items() 
                                if norm == normalized_name), normalized_name)
                
                # Initialize dictionary to store metrics by year
                metrics_by_year = {year: {} for year in years_range}
                
                # Process each year's data
                for year_data in years_data:
                    try:
                        tax_year = year_data.get('tax_year', 'Unknown')
                        if tax_year != 'Unknown':
                            tax_year = int(tax_year)
                            if tax_year in years_range:
                                metrics = year_data.get('financial_metrics', {})
                                
                                # Store metrics for this year
                                year_metrics = {}
                                
                                # Process metrics in the order of field_mapping
                                for display_col, field_name in self.field_mapping.items():
                                    value = metrics.get(field_name, None)
                                    year_metrics[display_col] = self.format_value(value, display_col)
                                
                                metrics_by_year[tax_year] = year_metrics
                    except Exception as e:
                        logger.error(f"Error processing year data for {display_name}: {str(e)}")
                        continue
                
                # Create rows for DataFrame using field_mapping order
                rows = []
                for metric_display_name in self.field_mapping.keys():
                    row = {'Metric': metric_display_name}
                    for year in years_range:
                        row[str(year)] = metrics_by_year[year].get(metric_display_name, None)
                    rows.append(row)
                
                # Create DataFrame if we have rows
                if rows:
                    df = pd.DataFrame(rows)
                    df.insert(0, 'Organization', display_name)
                    
                    # Store DataFrame with NTEE category
                    org_dfs[display_name] = {
                        'data': df,
                        'ntee_category': ntee_category
                    }
        
        return org_dfs

    def _normalize_org_name(self, name):
        """Normalize organization name for consistent matching"""
        if not name:
            return "Unknown"
            
        # Strip extra whitespace and convert to uppercase
        normalized = ' '.join(name.upper().split())
        
        # Remove common suffixes
        common_suffixes = [' INC', ' LLC', ' FOUNDATION', ' CORP', ' CORPORATION', ' LTD', ' INCORPORATED']
        for suffix in common_suffixes:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)]
        
        return normalized

    def write_to_excel(self, org_dfs):
        """Write data to Excel with each organization in its own sheet"""
        try:
            if not org_dfs:
                logger.error("No data to write to Excel")
                return False
                
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Create first sheet before removing default
                first_org = next(iter(org_dfs.items()))
                first_org_name, first_org_info = first_org
                first_df = first_org_info['data']
                first_ntee = first_org_info['ntee_category']
                first_sheet_name = self.clean_sheet_name(first_org_name)
                
                # Create and format first sheet
                worksheet = writer.book.create_sheet(first_sheet_name)
                worksheet['A1'] = f'NTEE Category: {first_ntee}'
                first_df.to_excel(writer, sheet_name=first_sheet_name, startrow=2, index=False)
                
                # Format first worksheet
                worksheet = writer.book[first_sheet_name]
                self._format_worksheet(worksheet, first_df)
                
                # Now safe to remove default sheet
                if 'Sheet' in writer.book.sheetnames:
                    std = writer.book['Sheet']
                    writer.book.remove(std)
                
                # Process remaining organizations
                remaining_orgs = list(org_dfs.items())[1:]
                for org_name, org_info in remaining_orgs:
                    df = org_info['data']
                    ntee_category = org_info['ntee_category']
                    sheet_name = self.clean_sheet_name(org_name)
                    
                    # Write the sheet
                    worksheet = writer.book.create_sheet(sheet_name)
                    worksheet['A1'] = f'NTEE Category: {ntee_category}'
                    df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
                    
                    # Format the worksheet
                    self._format_worksheet(worksheet, df)
            
            return True
            
        except Exception as e:
            logger.error(f"Error writing to Excel: {str(e)}")
            raise

    def _format_worksheet(self, worksheet, df):
        """Helper method to format worksheet"""
        # Format columns
        for idx, col in enumerate(df.columns):
            # Set column width
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx//26) + chr(65 + (idx % 26))
            worksheet.column_dimensions[col_letter].width = max_length + 2
            
            # Format numeric columns (year columns)
            if col.isdigit():
                for row in range(4, len(df) + 4):  # Skip header and NTEE category
                    cell = worksheet.cell(row=row, column=idx + 1)
                    
                    try:
                        if pd.notna(cell.value):
                            value = float(cell.value)
                            cell.value = round(value)  # Force integer for dollar amounts
                            cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'  # Accounting format
                    except (ValueError, TypeError):
                        continue
        
        # Format NTEE category row
        from openpyxl.styles import Font
        worksheet['A1'].font = Font(bold=True)


def main():
    print("Main function started!")
    
    # Initialize components
    parser = NonprofitPFParser()
    excel_handler = ExcelOutputHandlerPF(r'C:\Users\aronc\OneDrive\Documents\PushExcel_PF.xlsx')
    
    print("Components initialized successfully!")
    
    # Dictionary to store all org data by NTEE category
    all_org_data = {}
    
    print("Ready for input...")
    
    while True:
        # Get organization URL from user
        org_url = input("\nEnter the ProPublica organization URL (or type 'done' to finish): ")
        
        if org_url.lower() == 'done':
            break
            
        try:
            # Get NTEE category and XML links
            ntee_category, xml_links = parser.scraper.get_organization_links(org_url)
            
            # Initialize category if needed
            if ntee_category not in all_org_data:
                all_org_data[ntee_category] = {}
            
            logger.info(f"\nProcessing organization in category: {ntee_category}")
            
            for url in xml_links:
                try:
                    # Parse basic information
                    result = parser.process_url(url)
                    org_name = result['organization_name']
                    
                    # Initialize organization in data structure if needed
                    if org_name not in all_org_data[ntee_category]:
                        all_org_data[ntee_category][org_name] = []
                    
                    all_org_data[ntee_category][org_name].append(result)
                    logger.info(f"Successfully processed {url}")
                    
                except Exception as e:
                    logger.error(f"Error processing {url}: {str(e)}")
                    continue
                    
        except Exception as e:
            logger.error(f"Error processing organization: {str(e)}")
            continue
    
    # After all URLs are processed, write to Excel
    if all_org_data:
        category_dfs = excel_handler.consolidate_data(all_org_data)
        if excel_handler.write_to_excel(category_dfs):
            logger.info(f"\nSuccessfully wrote data to {excel_handler.output_path}")
        else:
            logger.error("Failed to write to Excel file")
    else:
        logger.error("No data was processed successfully")


if __name__ == "__main__":
    print("=== 990-PF Parser Starting ===")
    print("Initializing components...")
    main()