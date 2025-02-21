import xml.etree.ElementTree as ET
import pandas as pd
import requests
from io import BytesIO
import logging
from bs4 import BeautifulSoup
import os
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ScheduleHParser:
    def __init__(self, output_path):
        self.base_url = "https://projects.propublica.org"
        self.ns = {'irs': 'http://www.irs.gov/efile'}
        self.output_path = output_path
        
        # Field mapping for Excel column headers
        self.field_mapping = {
            # Key basic info
            'Organization': 'Organization',
            'Year': 'Year',
            
            # Financial Assistance fields
            'FinancialAssistanceAtCostTyp_TotalCommunityBenefitExpnsAmt': 'FinAssist Total Expense',
            'FinancialAssistanceAtCostTyp_DirectOffsettingRevenueAmt': 'FinAssist Offsetting Revenue',
            'FinancialAssistanceAtCostTyp_NetCommunityBenefitExpnsAmt': 'FinAssist Net Benefit',
            'FinancialAssistanceAtCostTyp_TotalExpensePct': 'FinAssist Expense Pct',

            # Unreimbursed Medicaid fields
            'UnreimbursedMedicaidGrp_TotalCommunityBenefitExpnsAmt': 'Medicaid Total Expense',
            'UnreimbursedMedicaidGrp_DirectOffsettingRevenueAmt': 'Medicaid Offsetting Revenue',
            'UnreimbursedMedicaidGrp_NetCommunityBenefitExpnsAmt': 'Medicaid Net Benefit',
            'UnreimbursedMedicaidGrp_TotalExpensePct': 'Medicaid Expense Pct',

            # Community Health Services fields
            'CommunityHealthServicesGrp_TotalCommunityBenefitExpnsAmt': 'CommHealth Total Expense',
            'CommunityHealthServicesGrp_DirectOffsettingRevenueAmt': 'CommHealth Offsetting Revenue',
            'CommunityHealthServicesGrp_NetCommunityBenefitExpnsAmt': 'CommHealth Net Benefit',
            'CommunityHealthServicesGrp_TotalExpensePct': 'CommHealth Expense Pct',

            # Health Professions Education fields
            'HealthProfessionsEducationGrp_TotalCommunityBenefitExpnsAmt': 'Education Total Expense',
            'HealthProfessionsEducationGrp_DirectOffsettingRevenueAmt': 'Education Offsetting Revenue',
            'HealthProfessionsEducationGrp_NetCommunityBenefitExpnsAmt': 'Education Net Benefit',
            'HealthProfessionsEducationGrp_TotalExpensePct': 'Education Expense Pct',

            # Research fields
            'ResearchGrp_TotalCommunityBenefitExpnsAmt': 'Research Total Expense',
            'ResearchGrp_DirectOffsettingRevenueAmt': 'Research Offsetting Revenue',
            'ResearchGrp_NetCommunityBenefitExpnsAmt': 'Research Net Benefit',
            'ResearchGrp_TotalExpensePct': 'Research Expense Pct',

            # Cash and In-Kind Contributions fields
            'CashAndInKindContributionsGrp_TotalCommunityBenefitExpnsAmt': 'Contributions Total Expense',
            'CashAndInKindContributionsGrp_DirectOffsettingRevenueAmt': 'Contributions Offsetting Revenue',
            'CashAndInKindContributionsGrp_NetCommunityBenefitExpnsAmt': 'Contributions Net Benefit',
            'CashAndInKindContributionsGrp_TotalExpensePct': 'Contributions Expense Pct',

            # Total Community Benefits fields
            'TotalCommunityBenefitsGrp_TotalCommunityBenefitExpnsAmt': 'Total Comm Benefits Expense',
            'TotalCommunityBenefitsGrp_DirectOffsettingRevenueAmt': 'Total Comm Benefits Offsetting',
            'TotalCommunityBenefitsGrp_NetCommunityBenefitExpnsAmt': 'Total Comm Benefits Net',
            'TotalCommunityBenefitsGrp_TotalExpensePct': 'Total Comm Benefits Pct',

            # Joint Ventures
            'JV1_Name': 'JV1 Name',
            'JV1_Activity': 'JV1 Activity',
            'JV1_OrgOwnership': 'JV1 Org Ownership',
            'JV1_PhysicianOwnership': 'JV1 Physician Ownership',
            
            'JV2_Name': 'JV2 Name',
            'JV2_Activity': 'JV2 Activity',
            'JV2_OrgOwnership': 'JV2 Org Ownership',
            'JV2_PhysicianOwnership': 'JV2 Physician Ownership',
            
            'JV3_Name': 'JV3 Name',
            'JV3_Activity': 'JV3 Activity',
            'JV3_OrgOwnership': 'JV3 Org Ownership',
            'JV3_PhysicianOwnership': 'JV3 Physician Ownership',
            
            'JV4_Name': 'JV4 Name',
            'JV4_Activity': 'JV4 Activity',
            'JV4_OrgOwnership': 'JV4 Org Ownership',
            'JV4_PhysicianOwnership': 'JV4 Physician Ownership',
            
            'JV5_Name': 'JV5 Name',
            'JV5_Activity': 'JV5 Activity',
            'JV5_OrgOwnership': 'JV5 Org Ownership',
            'JV5_PhysicianOwnership': 'JV5 Physician Ownership'
        }

    def get_xml_links(self, main_url):
        """Get XML download links from ProPublica page"""
        try:
            response = requests.get(main_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            xml_links = []
            for link in soup.find_all('a', class_='btn', target='_blank'):
                if 'XML' in link.text:
                    object_id = link['href'].split('object_id=')[1]
                    full_url = f"{self.base_url}/nonprofits/download-xml?object_id={object_id}"
                    xml_links.append(full_url)
            
            # Sort by object_id and take most recent 5
            xml_links.sort(reverse=True)
            return xml_links[:5]
            
        except Exception as e:
            logger.error(f"Error getting XML links: {e}")
            return []

    def get_org_name(self, root):
        """Extract organization name from XML"""
        for path in [
            './/irs:BusinessName/irs:BusinessNameLine1Txt',
            './/irs:ReturnHeader/irs:Filer/irs:BusinessName/irs:BusinessNameLine1Txt'
        ]:
            name = root.find(path, self.ns)
            if name is not None and name.text:
                return name.text
        return "Unknown Organization"

    def get_tax_year(self, root):
        """Extract tax year from XML"""
        tax_period = root.find('.//irs:TaxPeriodEndDt', self.ns)
        if tax_period is not None and tax_period.text:
            return tax_period.text[:4]  # Get just the year
        return "Unknown"
    
    def process_numeric_value(self, value, is_percentage=False):
        """Process numeric values, handling percentages appropriately"""
        try:
            if value is None or value.strip() == '':
                return None
                
            # Remove any commas and convert to float
            clean_value = float(value.replace(',', ''))
            
            if is_percentage:
                # If it's a percentage, leave as decimal (e.g., 0.0554 for 5.54%)
                return clean_value
            else:
                # For regular numbers, round to whole numbers
                return round(clean_value)
        except (ValueError, AttributeError):
            return None

    def extract_schedule_h(self, url):
        """Extract Schedule H data from a single XML file"""
        try:
            # Fetch and parse XML
            response = requests.get(url)
            tree = ET.parse(BytesIO(response.content))
            root = tree.getroot()
            
            # Get basic info
            org_name = self.get_org_name(root)
            tax_year = self.get_tax_year(root)
            
            # Initialize data dictionary
            data = {
                'Organization': org_name,
                'Year': tax_year
            }
            
            # List of all groups to extract
            groups = [
                'FinancialAssistanceAtCostTyp',
                'UnreimbursedMedicaidGrp',
                'UnreimbursedCostsGrp',
                'TotalFinancialAssistanceTyp',
                'CommunityHealthServicesGrp',
                'HealthProfessionsEducationGrp',
                'SubsidizedHealthServicesGrp',
                'ResearchGrp',
                'CashAndInKindContributionsGrp',
                'TotalOtherBenefitsGrp',
                'TotalCommunityBenefitsGrp',
                'PhysicalImprvAndHousingGrp',
                'EconomicDevelopmentGrp',
                'CommunitySupportGrp',
                'EnvironmentalImprovementsGrp',
                'LeadershipDevelopmentGrp',
                'CoalitionBuildingGrp',
                'HealthImprovementAdvocacyGrp',
                'WorkforceDevelopmentGrp',
                'OtherCommuntityBuildingActyGrp',
                'TotalCommuntityBuildingActyGrp'
            ]
            
            # Fields to extract for each group
            fields = [
                'TotalCommunityBenefitExpnsAmt',
                'DirectOffsettingRevenueAmt',
                'NetCommunityBenefitExpnsAmt',
                'TotalExpensePct'
            ]
            
            # Extract data for each group
            for group in groups:
                group_elem = root.find(f'.//irs:{group}', self.ns)
                if group_elem is not None:
                    for field in fields:
                        field_elem = group_elem.find(f'.//irs:{field}', self.ns)
                        if field_elem is not None and field_elem.text:
                            col_name = f"{group}_{field}"
                            # Process value based on whether it's a percentage field
                            is_percentage = field == 'TotalExpensePct'
                            data[col_name] = self.process_numeric_value(field_elem.text, is_percentage)
                        else:
                            data[f"{group}_{field}"] = None
            
            # Extract joint ventures
            ventures = []
            for idx, venture in enumerate(root.findall('.//irs:ManagementCoAndJntVenturesGrp', self.ns)):
                if idx >= 5:  # Limit to first 5 ventures
                    break
                    
                venture_data = {}
                # Get business name
                name_elem = venture.find('.//irs:BusinessNameLine1Txt', self.ns)
                if name_elem is not None:
                    data[f'JV{idx+1}_Name'] = name_elem.text
                
                # Get other venture fields
                fields = {
                    'PrimaryActivitiesTxt': f'JV{idx+1}_Activity',
                    'OrgProfitOrOwnershipPct': f'JV{idx+1}_OrgOwnership',
                    'PhysiciansProfitOrOwnershipPct': f'JV{idx+1}_PhysicianOwnership'
                }
                
                for field, col_name in fields.items():
                    elem = venture.find(f'.//irs:{field}', self.ns)
                    if elem is not None and elem.text:
                        data[col_name] = elem.text
                    else:
                        data[col_name] = None
            
            return data
            
        except Exception as e:
            logger.error(f"Error processing {url}: {e}")
            return None

    def process_organization(self, org_url):
        """Process all XML files for an organization"""
        xml_links = self.get_xml_links(org_url)
        all_data = []
        
        for url in xml_links:
            data = self.extract_schedule_h(url)
            if data:
                all_data.append(data)
                logger.info(f"Successfully processed {url}")
        
        return all_data

    def write_to_excel(self, data):
        """Write the extracted data to Excel with proper formatting"""
        try:
            if not data:
                logger.error("No data to write")
                return False
                
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Rename columns using the field mapping
            rename_dict = {col: self.field_mapping[col] for col in df.columns if col in self.field_mapping}
            df = df.rename(columns=rename_dict)
            
            # Separate joint ventures data
            jv_cols = [col for col in df.columns if col.startswith('JV')]
            jv_data = df[['Organization', 'Year'] + jv_cols].copy()
            
            # Remove JV columns from main DataFrame
            main_df = df.drop(columns=jv_cols)
            
            # Get all metric columns (excluding Organization and Year)
            metric_cols = [col for col in main_df.columns if col not in ['Organization', 'Year']]
            
            # Create pivoted DataFrame
            pivoted_df = pd.DataFrame()
            current_org = None
            
            for org in main_df['Organization'].unique():
                org_data = main_df[main_df['Organization'] == org]
                
                # Initialize metric names as index
                metrics_index = pd.Index(metric_cols, name='Metrics')
                
                # Create DataFrame for this organization
                org_pivoted = pd.DataFrame(index=metrics_index)
                
                # Add data for each year
                for _, row in org_data.iterrows():
                    year = str(row['Year'])
                    year_data = {metrics: row[metrics] for metrics in metric_cols}
                    org_pivoted[year] = pd.Series(year_data)
                
                # Add organization as a header
                org_pivoted = pd.concat([org_pivoted], keys=[org], names=['Organization'])
                
                # Append to main pivoted DataFrame
                pivoted_df = pd.concat([pivoted_df, org_pivoted])
                
                # Add joint ventures data for this organization
                org_jv = jv_data[jv_data['Organization'] == org].copy()
                if not org_jv.empty:
                    jv_pivoted = pd.DataFrame()
                    for _, jv_row in org_jv.iterrows():
                        year = str(jv_row['Year'])
                        jv_cols_subset = [col for col in jv_cols if not pd.isna(jv_row[col])]
                        for i in range(1, 6):  # For each JV
                            jv_data_row = {
                                'JV Name': jv_row.get(f'JV{i} Name', ''),
                                'Activity': jv_row.get(f'JV{i} Activity', ''),
                                'Org Ownership': jv_row.get(f'JV{i} Org Ownership', ''),
                                'Physician Ownership': jv_row.get(f'JV{i} Physician Ownership', '')
                            }
                            if any(jv_data_row.values()):  # Only add if there's data
                                jv_df = pd.DataFrame([jv_data_row], index=[f'Joint Venture {i}'])
                                jv_df.columns.name = year
                                jv_pivoted = pd.concat([jv_pivoted, jv_df])
                    
                    if not jv_pivoted.empty:
                        jv_pivoted = pd.concat([jv_pivoted], keys=[org], names=['Organization'])
                        pivoted_df = pd.concat([pivoted_df, jv_pivoted])
            
            # Write to Excel
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                pivoted_df.to_excel(writer, sheet_name='Sheet1')
                
                # Get worksheet
                ws = writer.sheets['Sheet1']
                
                # Format cells
                for row in range(2, ws.max_row + 1):
                    for col in range(3, ws.max_column + 1):  # Skip organization and metric columns
                        cell = ws.cell(row=row, column=col)
                        if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            metric_name = ws.cell(row=row, column=2).value  # Get metric name
                            if any(term in metric_name for term in ['Pct', 'Ownership']):
                                cell.number_format = '0.0000%'
                            else:
                                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            logger.info(f"Successfully wrote data to {self.output_path}")
            return True
        
        except Exception as e:
            logger.error(f"Error writing to Excel: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return False

def main():
    # Initialize parser with output path
    output_path = r'C:\Users\aronc\OneDrive\Documents\Betterformat.xlsx'
    
    # Check if the output file is currently open
    try:
        with open(output_path, 'a') as f:
            pass
    except PermissionError:
        print(f"Error: Please close the Excel file {output_path} before running the script.")
        return
        
    parser = ScheduleHParser(output_path)
    all_data = []
    
    while True:
        org_url = input("\nEnter the ProPublica organization URL (or type 'done' to finish): ").strip()
        
        if org_url.lower() == 'done':
            break
            
        if not org_url.startswith('http'):
            print("Please enter a valid URL starting with http:// or https://")
            continue
        
        # Process organization
        try:
            org_data = parser.process_organization(org_url)
            if org_data:
                all_data.extend(org_data)
                print(f"Successfully processed organization with {len(org_data)} years of data")
            else:
                print("No data was found for this organization")
        except Exception as e:
            print(f"Error processing organization: {str(e)}")
    
    # Write all data to Excel
    if all_data:
        print(f"\nProcessing {len(all_data)} records...")
        success = parser.write_to_excel(all_data)
        if success:
            print(f"\nExcel file has been created successfully at: {output_path}")
        else:
            print("\nFailed to create Excel file. Please check the error messages above.")
    else:
        print("\nNo data was processed successfully")
if __name__ == "__main__":
    main()